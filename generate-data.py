#!/usr/bin/env python3
"""
Liệt kê file hoặc thư mục con trong `original` (mặc định: ./original so với thư mục chứa script).
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def iter_files(root: Path, recursive: bool) -> list[Path]:
    """Trả về danh sách file (không gồm thư mục), sắp xếp theo đường dẫn."""
    if not root.is_dir():
        return []
    if recursive:
        paths = sorted(p for p in root.rglob("*") if p.is_file())
    else:
        paths = sorted(p for p in root.iterdir() if p.is_file())
    return paths


def iter_dirs(root: Path, recursive: bool) -> list[Path]:
    """Trả về danh sách thư mục con (không gồm file), sắp xếp theo đường dẫn."""
    if not root.is_dir():
        return []
    if recursive:
        paths = sorted(p for p in root.rglob("*") if p.is_dir())
    else:
        paths = sorted(p for p in root.iterdir() if p.is_dir())
    return paths


def relative_path_to_sheet_name(path_str: str) -> str:
    """Hai phần đầu của đường dẫn (vd. Myers\\Laptop\\...) -> tên sheet 'Myers - Laptop'."""
    norm = path_str.replace("\\", "/").strip("/")
    parts = Path(norm).parts
    if len(parts) >= 2:
        return f"{parts[0]} - {parts[1]}"
    if len(parts) == 1:
        return parts[0]
    return "Data"


def _sanitize_excel_sheet_name(name: str) -> str:
    """Excel giới hạn 31 ký tự; không dùng : \\ / ? * [ ]."""
    for ch in r":\/*?[]":
        name = name.replace(ch, "_")
    return name[:31] if len(name) > 31 else name


_EXCEL_EXPORT_COLUMNS = (
    "Path Folder",
    "Backup Name",
    "Timeline Name",
    "Short description",
    "Notes",
    "Status",
)


_THIN_BORDER = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(
    left=_THIN_BORDER,
    right=_THIN_BORDER,
    top=_THIN_BORDER,
    bottom=_THIN_BORDER,
)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGN = Alignment(vertical="top")
_DONE_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")

_EXCEL_COL_WIDTHS = {
    "A": 68,
    "B": 32,
    "C": 84,
    "D": 36,
    "E": 36,
    "F": 14,
}


def _apply_export_column_widths(ws) -> None:
    """Độ rộng cột export (A–F)."""
    for col, width in _EXCEL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _format_export_sheet(ws) -> None:
    """Cố định hàng tiêu đề, filter, viền, wrap cột dài, tô Status Done."""
    _apply_export_column_widths(ws)
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    last_col = get_column_letter(max_col)
    wrap_columns = {1, 3}  # Path Folder, Timeline Name

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _CELL_BORDER

    for row_idx in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = _CELL_BORDER
            cell.alignment = _WRAP_TOP if col in wrap_columns else _TOP_ALIGN
            if col == 6 and cell.value == "Done":
                cell.fill = _DONE_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{max_row}"


def _excel_export_row(folder: str, file_path: Path | None, result_json: list[dict]) -> dict[str, str]:
    backup_name = Path(file_path).name if file_path else ""
    timeline_name = ""
    status = ""
    for item in result_json:
        if item["folder"] == folder and item["backup_name"] == backup_name:
            timeline_name = item["timeline_name"]
            status = "Done"
            break
    # Replace \\ to / for Mac OS
    folder = f"{folder.replace("\\", "/")}"
    return {
        "Path Folder": folder,
        "Backup Name": backup_name,
        "Timeline Name": timeline_name,
        "Short description": "",
        "Notes": "",
        "Status": status,
    }


def _folder_data_to_rows_by_sheet(
    folder_data: list[dict],
    result_json: list[dict],
) -> dict[str, list[dict[str, str]]]:
    """Gom hàng export (giống Excel) theo tên sheet."""
    by_sheet: dict[str, list[dict[str, str]]] = defaultdict(list)
    last_folder_per_sheet: dict[str, str | None] = {}
    for item in folder_data:
        folder = item["folder"]
        files: list[Path] = item["files"]
        sheet = _sanitize_excel_sheet_name(relative_path_to_sheet_name(folder))
        last = last_folder_per_sheet.get(sheet)
        if files:
            for f in files:
                excel_row = _excel_export_row(folder, f, result_json)
                if excel_row["Backup Name"] != "":
                    by_sheet[sheet].append(excel_row)
                    last = folder
            last_folder_per_sheet[sheet] = last
        else:
            excel_row = _excel_export_row(folder, None, result_json)
            if excel_row["Backup Name"] != "":
                by_sheet[sheet].append(excel_row)
                last = folder
            last_folder_per_sheet[sheet] = folder
    return dict(by_sheet)


def export_folder_data_to_excel(
    folder_data: list[dict],
    output_dir: Path,
    result_json: list[dict],
    workbook_name: str = "data.xlsx",
) -> Path:
    """
    Ghi Excel vào `output_dir`: mỗi nhóm đường dẫn cùng hai phần đầu (Myers, Laptop)
    thành một sheet tên 'Myers - Laptop'.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / workbook_name

    by_sheet = _folder_data_to_rows_by_sheet(folder_data, result_json)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        if not by_sheet:
            pd.DataFrame(columns=list(_EXCEL_EXPORT_COLUMNS)).to_excel(
                writer, sheet_name="Data", index=False
            )
        else:
            for sheet_name, rows in by_sheet.items():
                df = pd.DataFrame(rows, columns=list(_EXCEL_EXPORT_COLUMNS))
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        for ws in writer.book.worksheets:
            _format_export_sheet(ws)

    return out_path

def read_json_file(file_path: Path) -> list[dict]:
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Lấy danh sách file hoặc thư mục con trong thư mục original."
    )
    parser.add_argument(
        "folder",
        nargs="?",
        default=None,
        help="Đường dẫn thư mục (mặc định: thư_mục_script/original)",
    )
    parser.add_argument(
        "--folders",
        "-f",
        action="store_true",
        help="Liệt kê thư mục con thay vì file",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="Đệ quy: gồm cả nội dung trong thư mục con",
    )
    parser.add_argument(
        "-a",
        "--absolute",
        action="store_true",
        help="In đường dẫn tuyệt đối",
    )
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    root = Path(args.folder).expanduser() if args.folder else script_dir / "original"
    root = root.resolve()
    result_json = read_json_file(script_dir / "output" / "restore_data.json")

    if not root.is_dir():
        print(f"Không tìm thấy thư mục: {root}", file=sys.stderr)
        print("Tạo thư mục `original` cạnh script hoặc truyền đường dẫn.", file=sys.stderr)
        return 1

    items = (
        iter_dirs(root, args.recursive)
        if args.folders
        else iter_files(root, args.recursive)
    )
    high_number_files = 0
    high_number_folder = ""
    numbers_files = 0
    folder_data = []
    for p in items:
        path = p.resolve() if args.absolute else p.relative_to(root)
        if len(path.parts) == 4:
            # Get list files in folder
            files = iter_files(p, args.recursive)
            if len(files) > high_number_files:
                high_number_files = len(files)
                high_number_folder = path
            numbers_files += len(files)
            folder_data.append({
                "folder": str(path),
                "files": files
            })
            continue
    print(f"High number files: {high_number_files} in folder: {high_number_folder}")
    print(f"Total files: {numbers_files}")
    out_dir = script_dir / "output"
    out = export_folder_data_to_excel(folder_data, out_dir, result_json)
    print(f"Đã xuất Excel: {out}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
